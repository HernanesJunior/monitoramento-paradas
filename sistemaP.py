import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import matplotlib
import numpy as np
import os
import json
import threading
import time
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from contextlib import contextmanager

ARQUIVO_CONFIG = 'config_app.json'
DIRETORIO_PADRAO = os.path.expanduser('~/Documents/ControleParadas')
ARQUIVO_TEMP = os.path.join(DIRETORIO_PADRAO, 'paradas_ativas.json')
PROCESSOS = ['Carregamento', 'Raku-Raku', 'Inspeção', 'Descarregamento']

MOTIVOS_POR_PROCESSO = {
    'Raku-Raku': [
        'Falta de bandeja', 'Ligar Cabine', 'Acumulo de bandeja', 'Sincronismos',
        'Alimentação', 'Engate', 'Falta de Dispositivos', 'Troca de Cor',
        'Falha de Pistola Automática', 'Atraso Aspiração', 'Ar-condicionado Cabine',
        'Troca de Pintor', 'Falha de Pistola Manual', 'Outros'
    ],
    'Carregamento': [
        'Tempo de Pausa intervalos', 'Falta de peça', 'Temperatura Estufa',
        'Mal Funcionamento Ebara', 'Falta de Espaço', 'Outros'
    ],
    'Inspeção': [
        'Falta de tinta', 'Lixo na Peça', 'Falta de Colaborador',
        'Fosfato na peça', 'Falha no Equipamento', 'Outros'
    ],
    'Descarregamento': [
        'HG/ELO Empenado para Robô', 'Queda de HG na Célula do Robô',
        'Sensor Robô Motoman', 'Sensor Robô Fanuc', 'Trans. Fora de Tempo para Robô',
        'Power Free Parou', 'Falta de Sincronismo', 'Falha no Pega do Robô',
        'Falha no Pega Robô Motoman', 'Acumulo de Bandeja no Robô',
        'Trans. LSA Parou', 'Falha de Pistão/LS Power Free',
        'Chassi Descendo Moldado', 'Outros'
    ]
}

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

class AplicativoMobile(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Controle de Paradas")
        self.geometry("360x540")
        # permitir redimensionamento para melhor usabilidade em telas maiores
        self.minsize(360, 540)
        self.maxsize(800, 800)
        self.resizable(True, True)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.nome_funcionario = ctk.StringVar()
        self.processo_selecionado = ""
        self.paradas_em_andamento = []
        self.config = self.carregar_config()
        self.carregar_paradas_ativas()
        self.thread_salvamento_ativa = False # Flag para controlar a thread de salvamento

        self.verificar_diretorio()
        self.criar_estilos()
        self.criar_tela_login()
        self.protocol("WM_DELETE_WINDOW", self.ao_fechar)

    @contextmanager
    def abrir_planilha(self, arquivo):
        wb = None
        try:
            if os.path.exists(arquivo):
                wb = load_workbook(arquivo)
            else:
                wb = Workbook()
                wb.remove(wb.active)
            yield wb
        except Exception as e:
            raise RuntimeError(f"Erro ao acessar planilha: {str(e)}")
        finally:
            if wb:
                try:
                    wb.save(arquivo)
                    wb.close()
                except:
                    pass

    def carregar_config(self):
        # garante que a configuração sempre contenha as chaves esperadas
        defaults = {
            'diretorio': DIRETORIO_PADRAO,
            'remember_me': False,
            'last_user': ''
        }
        try:
            with open(ARQUIVO_CONFIG, 'r') as f:
                cfg = json.load(f)
                # mescla defaults com os valores existentes
                for k, v in defaults.items():
                    cfg.setdefault(k, v)
                return cfg
        except (FileNotFoundError, json.JSONDecodeError):
            return defaults

    def salvar_config(self):
        with open(ARQUIVO_CONFIG, 'w') as f:
            json.dump(self.config, f)

    def carregar_paradas_ativas(self):
        try:
            with open(ARQUIVO_TEMP, 'r') as f:
                self.paradas_em_andamento = json.load(f)
                for parada in self.paradas_em_andamento:
                    parada['inicio'] = datetime.strptime(parada['inicio'], '%Y-%m-%d %H:%M:%S')
        except (FileNotFoundError, json.JSONDecodeError):
            pass

    def salvar_paradas_ativas(self):
        temp = []
        for parada in self.paradas_em_andamento:
            temp_parada = parada.copy()
            temp_parada['inicio'] = temp_parada['inicio'].strftime('%Y-%m-%d %H:%M:%S')
            temp.append(temp_parada)
        with open(ARQUIVO_TEMP, 'w') as f:
            json.dump(temp, f)

    def verificar_diretorio(self):
        os.makedirs(self.config['diretorio'], exist_ok=True)

    def caminho_arquivo(self, nome_arquivo):
        return os.path.join(self.config['diretorio'], nome_arquivo)

    def selecionar_diretorio(self):
        novo_dir = filedialog.askdirectory(
            initialdir=self.config['diretorio'],
            title="Selecione o diretório para salvar os arquivos"
        )
        if novo_dir:
            self.config['diretorio'] = novo_dir
            self.salvar_config()
            self.verificar_diretorio()
            messagebox.showinfo("Sucesso", f"Diretório alterado para:\n{novo_dir}")

    def criar_estilos(self):
        self.font_titulo = ctk.CTkFont(family="Helvetica", size=20, weight="bold")
        self.font_texto = ctk.CTkFont(family="Helvetica", size=14)
        self.font_botao = ctk.CTkFont(family="Helvetica", size=14)
        self.cor_principal = "#3498db"  # Cor principal dos botões
        self.cor_fundo = "#f5f5f5"      # Cor de fundo clara
        self.cor_card = "#222222"        # Cor dos "cards"
        self.cor_texto_principal = "#ffffff" # Cor do texto principal (branco para fundo escuro)
        self.cor_botao_texto = "white"  # Cor do texto dos botões principais
        self.cor_motivo_botao = "#444444" # Cor dos botões de motivo (mais escuro)
        self.cor_motivo_texto = "white" # Cor do texto dos botões de motivo (branco)
        self.cor_parada_ativa = "#404040" # Cor de fundo para paradas ativas (mais escuro)
        self.cor_borda_card = "#555555" # Cor da borda dos cards (mais clara que o fundo)
        self.cor_borda_parada_ativa = "#666666" # Cor da borda para paradas ativas (mais clara)

    def limpar_tela(self):
        for widget in self.winfo_children():
            widget.destroy()

    def criar_tela_login(self):
        self.limpar_tela()

        ctk.CTkLabel(self, text="Controle de Paradas", font=self.font_titulo, text_color="#d3d3d3").pack(pady=20)

        ctk.CTkLabel(self, text="Operador:", font=self.font_texto, text_color="#d3d3d3").pack(pady=10)

        self.entry_nome_operador = ctk.CTkEntry(self, placeholder_text='Digite seu usuário:')
        self.entry_nome_operador.pack(pady=10, ipady=8, padx=20, fill='x')
        self.entry_nome_operador.focus()

        ctk.CTkLabel(self, text='Senha', font=self.font_texto, text_color="#d3d3d3").pack(pady=10)

        self.entry_senha_funcionario = ctk.CTkEntry(self, placeholder_text='Digite sua senha:', show="*")
        self.entry_senha_funcionario.pack(pady=10, ipady=8, padx=20, fill='x')

        # checkbox para mostrar/ocultar senha
        self.show_password_var = ctk.BooleanVar(value=False)
        chk_show = ctk.CTkCheckBox(self, text='Mostrar senha', variable=self.show_password_var, command=self._toggle_password_visibility)
        chk_show.pack(pady=5)

        # lembrar senha / usuário
        self.remember_var = ctk.BooleanVar(value=self.config.get('remember_me', False))
        checkbox = ctk.CTkCheckBox(self, text='Lembrar usuário', variable=self.remember_var)
        checkbox.pack(padx=10, pady=10)

        ctk.CTkButton(self, text="Entrar", command=self.verificar_nome, height=50, font=self.font_botao).pack(pady=20, padx=20, fill='x')
        self.label_resultado_login = ctk.CTkLabel(self, text='')
        self.label_resultado_login.pack(pady=5)

        # preencher campos se havia usuário lembrado
        if self.config.get('remember_me'):
            self.entry_nome_operador.insert(0, self.config.get('last_user', ''))

    def _toggle_password_visibility(self):
        # alterna entre mostrar e ocultar o conteúdo da senha
        if self.show_password_var.get():
            self.entry_senha_funcionario.configure(show="")
        else:
            self.entry_senha_funcionario.configure(show="*")

    def verificar_nome(self):
        nome = self.entry_nome_operador.get().strip()
        senha = self.entry_senha_funcionario.get()
        if nome == 'operador' and senha == '123':
            self.label_resultado_login.configure(text="Login bem-sucedido!", text_color="green")
            self.nome_funcionario.set(nome) # Define o nome do funcionário
            # salvar preferências de usuário
            if self.remember_var.get():
                self.config['remember_me'] = True
                self.config['last_user'] = nome
            else:
                self.config['remember_me'] = False
                self.config['last_user'] = ''
            self.salvar_config()
            self.after(1500, self.criar_tela_processos) # Redireciona após um breve delay
        else:
            self.label_resultado_login.configure(text="Nome ou senha incorretos!", text_color="red")

    def criar_tela_processos(self):
        self.limpar_tela()

        # container rolável para processar muitos botões se necessário
        container = ctk.CTkFrame(self)
        container.pack(expand=True, fill='both', padx=15, pady=15)
        canvas = ctk.CTkCanvas(container, highlightthickness=0)
        scrollbar = ctk.CTkScrollbar(container, orientation='vertical', command=canvas.yview)
        scroll_frame = ctk.CTkFrame(canvas)
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        scroll_frame.columnconfigure(0, weight=1)
        scroll_frame.columnconfigure(1, weight=1)

        for i, processo in enumerate(PROCESSOS):
            btn = ctk.CTkButton(scroll_frame, text=processo, font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=lambda p=processo: self.selecionar_processo(p))
            btn.grid(row=i//2, column=i%2, padx=5, pady=5, sticky='nsew')
            btn.configure(width=150, height=50)

        self.criar_status_bar()
        self.atualizar_status_bar()

        rodape = ctk.CTkFrame(self, fg_color=self.cor_card)
        rodape.pack(fill='x', padx=15, pady=15) # Aumentei padx e pady
        rodape.columnconfigure(0, weight=1)
        rodape.columnconfigure(1, weight=1)
        rodape.columnconfigure(2, weight=1)
        rodape.columnconfigure(3, weight=1)

        ctk.CTkButton(rodape, text="Paradas Ativas", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.mostrar_paradas_ativas).grid(row=0, column=0, padx=5, pady=5, sticky='ew')
        ctk.CTkButton(rodape, text="Histórico", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.mostrar_historico).grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        ctk.CTkButton(rodape, text="Sair", font=self.font_botao, text_color=self.cor_botao_texto, fg_color='DarkRed', hover_color="#2980b9", command=self.ao_fechar).grid(row=0, column=2, padx=5, pady=5, sticky='ew')
        ctk.CTkButton(rodape, text="⚙", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.mostrar_configuracoes, width=50).grid(row=0, column=3, padx=5, pady=5, sticky='e')

    def selecionar_processo(self, processo):
        self.processo_selecionado = processo
        self.criar_tela_motivos()

    def criar_tela_motivos(self):
        self.limpar_tela()

        cabecalho = ctk.CTkFrame(self, fg_color=self.cor_card)
        cabecalho.pack(fill='x', padx=10, pady=10)
        ctk.CTkButton(cabecalho, text="← Voltar", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.criar_tela_processos).pack(side='left', padx=5) # Adicionei padx
        ctk.CTkLabel(cabecalho, text=self.processo_selecionado, font=self.font_texto, text_color=self.cor_texto_principal).pack(side='left', padx=10)

        container = ctk.CTkFrame(self)
        container.pack(expand=True, fill='both', padx=15, pady=15) # Aumentei padx e pady
        container.columnconfigure(0, weight=1)

        motivos = MOTIVOS_POR_PROCESSO.get(self.processo_selecionado, [])

        for motivo in motivos:
            btn = ctk.CTkButton(container, text=motivo, font=self.font_texto, text_color=self.cor_motivo_texto, fg_color=self.cor_motivo_botao, hover_color="#cccccc", border_width=1, border_color=self.cor_borda_card, command=lambda m=motivo: self.registrar_parada(m))
            btn.pack(fill='x', pady=3, padx=5) # Adicionei padx e um pequeno pady

        if 'Outros' in motivos:
            self.entry_outros = ctk.CTkEntry(container, font=self.font_texto)
            self.entry_outros.pack(fill='x', pady=5, padx=5)
            self.entry_outros.insert(0, "Descreva o motivo...")
            self.entry_outros.bind("<FocusIn>", lambda e: self.entry_outros.delete(0, 'end'))

    def registrar_parada(self, motivo):
        if motivo == 'Outros':
            motivo = self.entry_outros.get().strip()
            if not motivo or motivo == "Descreva o motivo...":
                messagebox.showwarning("Atenção", "Descreva o motivo da parada!")
                return

        nova_parada = {
            'funcionario': self.nome_funcionario.get(),
            'processo': self.processo_selecionado,
            'motivo': motivo,
            'inicio': datetime.now(),
            'fim': None,
            'duracao': None
        }

        self.paradas_em_andamento.append(nova_parada)
        self.criar_status_bar()
        self.atualizar_status_bar()
        self.mostrar_paradas_ativas()

    def mostrar_paradas_ativas(self):
        self.limpar_tela()
        self.criar_status_bar()
        self.atualizar_status_bar()

        cabecalho = ctk.CTkFrame(self, fg_color=self.cor_card)
        cabecalho.pack(fill='x', padx=10, pady=10)
        ctk.CTkButton(cabecalho, text="← Voltar", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.criar_tela_processos).pack(side='left', padx=5) # Adicionei padx
        ctk.CTkLabel(cabecalho, text="Paradas Ativas", font=self.font_texto, text_color=self.cor_texto_principal).pack(side='left', padx=10)

        if not self.paradas_em_andamento:
            ctk.CTkLabel(self, text="Nenhuma parada ativa.", font=self.font_texto, text_color=self.cor_texto_principal).pack(pady=20)
            return

        container_paradas = ctk.CTkFrame(self)
        container_paradas.pack(expand=True, fill='both')
        container_paradas.columnconfigure(0, weight=1)
        container_paradas.rowconfigure(0, weight=1)

        canvas = ctk.CTkCanvas(container_paradas, highlightthickness=0)
        barra_rolagem = ctk.CTkScrollbar(container_paradas, orientation="vertical", command=canvas.yview)
        frame_rolavel = ctk.CTkFrame(canvas)

        frame_rolavel.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame_rolavel, anchor="nw")
        canvas.configure(yscrollcommand=barra_rolagem.set)

        canvas.pack(side="left", fill="both", expand=True)
        barra_rolagem.pack(side="right", fill="y")

        for parada in self.paradas_em_andamento:
            frame_parada = ctk.CTkFrame(frame_rolavel, fg_color=self.cor_parada_ativa, border_color=self.cor_borda_parada_ativa, border_width=1)
            frame_parada.pack(fill='x', pady=5, padx=5)
            frame_parada.columnconfigure(0, weight=1)
            frame_parada.columnconfigure(1, weight=0)

            inicio_formatado = parada['inicio'].strftime('%H:%M:%S')
            label_info = ctk.CTkLabel(frame_parada, text=f"{parada['processo']} - {parada['motivo']}\nInício: {inicio_formatado}", font=self.font_texto, text_color=self.cor_texto_principal, anchor='w')
            label_info.grid(row=0, column=0, padx=10, pady=5, sticky='ew')

            btn_finalizar = ctk.CTkButton(frame_parada, text="Finalizar", font=self.font_botao, text_color=self.cor_botao_texto, fg_color='DarkGreen', hover_color="#2e8b57", command=lambda p=parada: self.finalizar_parada(p))
            btn_finalizar.grid(row=0, column=1, padx=10, pady=5, sticky='e')

    def finalizar_parada(self, parada):
        fim = datetime.now()
        duracao = (fim - parada['inicio']).total_seconds() / 60
        parada['fim'] = fim
        parada['duracao'] = duracao

        self.salvar_parada_historico(parada)
        self.paradas_em_andamento.remove(parada)
        self.salvar_paradas_ativas()
        self.mostrar_paradas_ativas()
        # atualiza barra de status após finalizar
        self.atualizar_status_bar()

    def salvar_parada_historico(self, parada):
        arquivo_excel = self.caminho_arquivo('paradas.xlsx')
        try:
            with self.abrir_planilha(arquivo_excel) as wb:
                if 'Paradas' not in wb.sheetnames:
                    ws = wb.create_sheet('Paradas')
                    ws.append(['Data', 'Processo', 'Funcionário', 'Motivo', 'Início', 'Fim', 'Duração (min)'])
                    ws.column_dimensions['A'].width = 15
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 20
                    ws.column_dimensions['D'].width = 30
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15
                    cabecalho = ws['A1:G1']
                    for cell in cabecalho[0]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(bottom=Side(style='thin'))
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                else:
                    ws = wb['Paradas']

                ws.append([
                    parada['inicio'].strftime('%Y-%m-%d'),
                    parada['processo'],
                    parada['funcionario'],
                    parada['motivo'],
                    parada['inicio'].strftime('%H:%M:%S'),
                    parada['fim'].strftime('%H:%M:%S'),
                    parada['duracao']
                ])

                ultima_linha = ws.max_row
                intervalo_tempo = ws[f'E{ultima_linha}:F{ultima_linha}']
                for cell in intervalo_tempo[0]:
                    cell.alignment = Alignment(horizontal='center')

                duracao_cell = ws[f'G{ultima_linha}']
                duracao_cell.number_format = '0.00'
                duracao_cell.alignment = Alignment(horizontal='right')

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar histórico: {e}")

    def mostrar_historico(self):
        self.limpar_tela()
        self.criar_status_bar()
        self.atualizar_status_bar()

        tabview = ctk.CTkTabview(self)
        tabview.pack(expand=True, fill='both', padx=10, pady=10)

        # Aba de Histórico
        tabview.add("Histórico")
        container_historico = ctk.CTkFrame(tabview.tab("Histórico"))
        container_historico.pack(expand=True, fill='both')
        container_historico.columnconfigure(0, weight=1)

        canvas_historico = ctk.CTkCanvas(container_historico, highlightthickness=0)
        barra_rolagem_historico = ctk.CTkScrollbar(container_historico, orientation="vertical", command=canvas_historico.yview)
        frame_rolavel_historico = ctk.CTkFrame(canvas_historico)
        frame_rolavel_historico.columnconfigure(0, weight=1)
        frame_rolavel_historico.bind("<Configure>", lambda e: canvas_historico.configure(scrollregion=canvas_historico.bbox("all")))
        canvas_historico.create_window((0, 0), window=frame_rolavel_historico, anchor="nw")
        canvas_historico.configure(yscrollcommand=barra_rolagem_historico.set)
        canvas_historico.pack(side="left", fill="both", expand=True)
        barra_rolagem_historico.pack(side="right", fill="y")
        self.atualizar_historico(frame_rolavel_historico)

        # Aba de Gráficos com rolagem na aba inteira e gráficos na vertical
        tab_graficos = tabview.add("Gráficos")
        canvas_graficos_aba = ctk.CTkCanvas(tab_graficos, highlightthickness=0)
        barra_rolagem_graficos_aba = ctk.CTkScrollbar(tab_graficos, orientation="vertical", command=canvas_graficos_aba.yview)
        self.frame_graficos = ctk.CTkFrame(canvas_graficos_aba)
        self.frame_graficos.columnconfigure(0, weight=1)
        self.frame_graficos.bind("<Configure>", lambda e: canvas_graficos_aba.configure(scrollregion=canvas_graficos_aba.bbox("all")))
        canvas_graficos_aba.create_window((0, 0), window=self.frame_graficos, anchor="nw")
        canvas_graficos_aba.configure(yscrollcommand=barra_rolagem_graficos_aba.set)
        canvas_graficos_aba.pack(side="left", fill="both", expand=True)
        barra_rolagem_graficos_aba.pack(side="right", fill="y")
        self.gerar_graficos_historicos(self.frame_graficos) # Gerar os gráficos dentro do self.frame_graficos

        # Cabeçalho e botões (sem alterações para esta questão)
        cabecalho = ctk.CTkFrame(self, fg_color=self.cor_card)
        cabecalho.pack(fill='x', padx=10, pady=10)
        ctk.CTkButton(cabecalho, text="← Voltar", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.criar_tela_processos).pack(side='left')
        ctk.CTkLabel(cabecalho, text="Histórico e Gráficos", font=self.font_texto, text_color=self.cor_texto_principal).pack(side='left', padx=10)

        frame_botoes_rodape = ctk.CTkFrame(cabecalho)
        frame_botoes_rodape.pack(side='right')
        ctk.CTkButton(frame_botoes_rodape, text="🔄", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.atualizar_historico).pack(side='left', padx=2)
        ctk.CTkButton(frame_botoes_rodape, text="🗑️", font=self.font_botao, text_color=self.cor_botao_texto, fg_color=self.cor_principal, hover_color="#2980b9", command=self.apagar_relatorio).pack(side='left', padx=2)

    def gerar_graficos_historicos(self, container):
        for widget in container.winfo_children():
            widget.destroy()
        container.columnconfigure(0, weight=1) # Garante que o conteúdo se expande horizontalmente

        try:
            arquivo_excel = self.caminho_arquivo('paradas.xlsx')
            with self.abrir_planilha(arquivo_excel) as wb:
                if 'Paradas' not in wb.sheetnames:
                    ctk.CTkLabel(container, text="Nenhum dado de parada registrado.", text_color='red').pack(pady=20)
                    return
                ws_paradas = wb['Paradas']

                dados_processos = {}
                dados_motivos = {}
                total_tempo = 0

                for row in ws_paradas.iter_rows(min_row=2, max_col=7, values_only=True):
                    if row and row[0] != 'TOTAL' and row[1] and row[3] and row[6]:
                        processo = row[1]
                        motivo = row[3]
                        duracao = row[6] or 0
                        total_tempo += duracao
                        dados_processos[processo] = dados_processos.get(processo, 0) + duracao
                        dados_motivos[motivo] = dados_motivos.get(motivo, 0) + duracao

                frame_processos = ctk.CTkFrame(container)
                frame_processos.pack(fill='x', pady=5) # Use pack para empilhar verticalmente
                self.criar_grafico_barras_porcentagem(
                    frame_processos,
                    dados_processos,
                    "Tempo de Parada por Processo",
                    total_tempo
                )

                frame_motivos = ctk.CTkFrame(container)
                frame_motivos.pack(fill='x', pady=5) # Use pack para empilhar verticalmente
                self.criar_grafico_barras_porcentagem(
                    frame_motivos,
                    dados_motivos,
                    "Tempo de Parada por Motivo",
                    total_tempo
                )

        except FileNotFoundError:
            ctk.CTkLabel(container, text="Arquivo de histórico não encontrado.", text_color='red').pack(pady=20)
        except KeyError:
            ctk.CTkLabel(container, text="Dados insuficientes para gerar gráficos.", text_color='red').pack(pady=20)
        except Exception as e:
            ctk.CTkLabel(container, text=f"Erro ao gerar gráficos: {str(e)}", text_color='red').pack()

    def criar_grafico_barras_porcentagem(self, container, dados, titulo, total_tempo):
        try:
            if not dados:
                raise ValueError("Nenhum dado disponível")

            fig = plt.figure(figsize=(8, 6), dpi=100, facecolor=self.cor_fundo)
            ax = fig.add_subplot(111)

            labels = list(dados.keys())
            tempos = list(dados.values())
            porcentagens = [(t / total_tempo) * 100 if total_tempo > 0 else 0 for t in tempos]

            num_labels = len(labels)
            cores = plt.cm.viridis(np.linspace(0, 1, num_labels)) # Usar o mapa de cores 'viridis'

            x = range(num_labels)
            bars = ax.bar(x, tempos, color=cores, align='center') # Aplicar as cores

            ax.set_ylabel("Tempo Total (minutos)", fontsize=12)
            ax.set_xlabel("Processo/Motivo", fontsize=12)
            ax.set_title(titulo, fontsize=14)

            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=10)

            ax.tick_params(axis='y', labelsize=10)
            ax.grid(axis='y', linestyle='--')

            for bar, porcentagem in zip(bars, porcentagens):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, height + 0.5, f'{porcentagem:.1f}%', ha='center', va='bottom', fontsize=9)

            plt.tight_layout(pad=1.5, w_pad=2.0, h_pad=2.0)

            canvas = FigureCanvasTkAgg(fig, master=container)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

        except ValueError as ve:
            ctk.CTkLabel(container, text=f"Erro no gráfico ({titulo}): {str(ve)}", text_color='red').pack()
        except Exception as e:
            ctk.CTkLabel(container, text=f"Erro no gráfico ({titulo}): {str(e)}", text_color='red').pack()

    def atualizar_historico(self, container):
        if container:
            for widget in container.winfo_children():
                widget.destroy()

            try:
                arquivo_excel = self.caminho_arquivo('paradas.xlsx')
                if not os.path.exists(arquivo_excel):
                    ctk.CTkLabel(container, text="Nenhum registro encontrado", font=self.font_texto, text_color=self.cor_texto_principal).pack()
                    return

                with self.abrir_planilha(arquivo_excel) as wb:
                    if 'Paradas' not in wb.sheetnames:
                        ctk.CTkLabel(container, text="Nenhum registro de paradas encontrado.", font=self.font_texto, text_color=self.cor_texto_principal).pack()
                        return
                    ws_paradas = wb['Paradas']

                    for row in ws_paradas.iter_rows(min_row=2, values_only=True):
                        if row and row[0] == 'TOTAL':
                            continue

                        frame = ctk.CTkFrame(container, fg_color=self.cor_card, border_color=self.cor_borda_card, border_width=1)
                        frame.pack(fill='x', pady=8)
                        frame.columnconfigure(0, weight=1)

                        texto = (f"Data: {row[0]}\n"
                                 f"Processo: {row[1]}\n"
                                 f"Motivo: {row[3]}\n"
                                 f"Duração: {row[6]:.2f} min ({row[4]} - {row[5]})")

                        ctk.CTkLabel(frame, text=texto, font=self.font_texto, text_color=self.cor_texto_principal, anchor='w', padx=20).pack(side='left', fill='x', expand=True, pady=20)

            except FileNotFoundError:
                ctk.CTkLabel(container, text="Nenhum registro encontrado", font=self.font_texto, text_color=self.cor_texto_principal).pack()
            except KeyError:
                ctk.CTkLabel(container, text="Estrutura do arquivo de histórico inválida.", font=self.font_texto, text_color='red').pack()
    def apagar_relatorio(self):
        arquivo_excel = self.caminho_arquivo('paradas.xlsx')
        if os.path.exists(arquivo_excel):
            confirmacao = messagebox.askyesno("Atenção", "Deseja apagar todo o histórico de paradas?")
            if confirmacao:
                try:
                    os.remove(arquivo_excel)
                    messagebox.showinfo("Sucesso", "Histórico de paradas apagado!")
                    # Recarregar a tela para refletir a mudança nas abas
                    self.mostrar_historico()
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao apagar o arquivo: {e}")
        else:
            messagebox.showinfo("Informação", "Nenhum histórico de paradas encontrado.")

    def mostrar_configuracoes(self):
        tela_config = ctk.CTkToplevel(self)
        tela_config.title("Configurações")
        tela_config.geometry("300x150")
        tela_config.resizable(False, False)

        ctk.CTkLabel(tela_config, text="Diretório de Salvamento:", font=self.font_texto).pack(pady=10)
        entry_diretorio = ctk.CTkEntry(tela_config, width=200)
        entry_diretorio.insert(0, self.config['diretorio'])
        entry_diretorio.configure(state='disabled')
        entry_diretorio.pack(pady=5)

        btn_selecionar_diretorio = ctk.CTkButton(tela_config, text="Alterar Diretório", command=self.selecionar_diretorio)
        btn_selecionar_diretorio.pack(pady=10)

    def criar_status_bar(self):
        # adiciona barra de status na parte inferior da janela
        if hasattr(self, 'status_bar') and self.status_bar.winfo_exists():
            return
        self.status_bar = ctk.CTkLabel(self, text="", font=self.font_texto, fg_color=self.cor_card, anchor='w')
        self.status_bar.pack(side='bottom', fill='x')

    def atualizar_status_bar(self):
        # mostra número de paradas ativas e usuário logado
        try:
            count = len(self.paradas_em_andamento)
            user = self.nome_funcionario.get()
            self.status_bar.configure(text=f"Paradas ativas: {count} • Usuário: {user}")
        except Exception:
            pass

    def ao_fechar(self):
        self.salvar_paradas_ativas()
        self.destroy()

if __name__ == "__main__":
    app = AplicativoMobile()
    app.mainloop()

    #  ;-)